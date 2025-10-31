from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from .models import Personal
from .forms import PersonalForm

def personal_list(request):
    proyecto = request.GET.get('proyecto')
    estado = request.GET.get('estado')
    qs = Personal.objects.all().order_by('-id')
    if proyecto: qs = qs.filter(proyecto=proyecto)
    if estado: qs = qs.filter(estado=estado)
    return render(request, 'personal/list.html', {
        'personal_list': qs, 'f_proyecto': proyecto or '', 'f_estado': estado or ''
    })

def personal_detail(request, pk):
    obj = get_object_or_404(Personal, pk=pk)
    return render(request, 'personal/detail.html', {'obj': obj})

def personal_create(request):
    if request.method == 'POST':
        form = PersonalForm(request.POST)
        if form.is_valid():
            form.save(); messages.success(request, 'Personal creado correctamente.')
            return redirect('personal:list')
    else:
        form = PersonalForm()
    return render(request, 'personal/form.html', {'form': form, 'title': 'Nuevo Personal'})

def personal_update(request, pk):
    obj = get_object_or_404(Personal, pk=pk)
    if request.method == 'POST':
        form = PersonalForm(request.POST, instance=obj)
        if form.is_valid():
            form.save(); messages.success(request, 'Personal actualizado.')
            return redirect('personal:list')
    else:
        form = PersonalForm(instance=obj)
    return render(request, 'personal/form.html', {'form': form, 'title': 'Editar Personal'})

def personal_delete(request, pk):
    obj = get_object_or_404(Personal, pk=pk)
    if request.method == 'POST':
        obj.delete(); messages.success(request, 'Registro eliminado.')
        return redirect('personal:list')
    return render(request, 'personal/confirm_delete.html', {'obj': obj})

def exportar_csv(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.utils.encoding import smart_str
    proyecto = request.GET.get('proyecto'); estado = request.GET.get('estado')
    qs = Personal.objects.all().order_by('id')
    if proyecto: qs = qs.filter(proyecto=proyecto)
    if estado: qs = qs.filter(estado=estado)
    wb = Workbook(); ws = wb.active; ws.title = "Personal"
    headers = ["ID","Nombre","DPI","Teléfono","Proyecto","Estado","Fecha Ingreso","Observaciones"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for c in range(1,len(headers)+1):
        cell = ws.cell(row=1, column=c); cell.font = head_font; cell.fill = head_fill; cell.alignment = Alignment(horizontal="center")
    for p in qs:
        ws.append([p.id, smart_str(p.nombre), smart_str(p.dpi), smart_str(p.telefono),
                   p.get_proyecto_display(), p.get_estado_display(), p.fecha_ingreso.strftime("%Y-%m-%d"),
                   smart_str(p.observaciones or "")])
    for col in ws.columns:
        max_len = 0; col_letter = col[0].column_letter
        for cell in col:
            try: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = max_len + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=personal_export.xlsx'
    wb.save(response); return response
